# -*- coding: utf-8 -*-
import os
import openpyxl
import datetime
import sys

class Spliter(object):
    def Run(self, targetFile, maxRow = 0, outputDir = None, outputFileName = None):
        rowOneCellNames = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1"]
        inputBook = openpyxl.load_workbook(targetFile, read_only=True)
        inputSheet = inputBook[inputBook.sheetnames[0]]
        if maxRow != 0:
            inputSheet.max_row = maxRow
        columsName = []
        for i in rowOneCellNames:
            columsName.append(inputSheet[i].value)
        
        print("Begin loading input file : %s ..." % targetFile)
        rowIndex = 1
        maxIndex = inputSheet.max_row
        allData = {}
        isFirstRow = True
        for row in inputSheet:

            sys.stdout.write("processing index %d/%d \r" %(rowIndex, maxIndex))
            sys.stdout.flush()
            rowIndex += 1

            if isFirstRow:
                isFirstRow = False
                continue
            rowData = None
            rowID = ""
            oneRow = []
            for cell in row:
                if rowData == None:
                    rowID = cell.value
                    if rowID == None:
                        break
                    rowData = allData.get(rowID)
                    if rowData == None:
                        rowData = []
                        allData[rowID] = rowData
                oneRow.append(cell.value)
            if rowData != None:
                rowData.append(oneRow)
        print("\nprocess Done!")
        
        print("Begin writing files...")
        if outputDir != None:
            os.makedirs(outputDir, exist_ok=True)
        fileIndex = 1
        maxFile = len(allData)
        for fileName in allData:
            if fileName == None:
                continue
            sys.stdout.write("processing file %d/%d \r" %(fileIndex, maxFile))
            sys.stdout.flush()
            fileIndex += 1

            writeBook = openpyxl.Workbook(write_only=True)
            writeSheet = writeBook.create_sheet(title=fileName)
            writeSheet.append(columsName)
            for row in allData[fileName]:
                writeSheet.append(row)
            outputPath = fileName+".xlsx"
            if outputDir != None:
                outputPath = outputDir + "\\" + outputPath
            writeBook.save(outputPath)
        print("\nWrite Files Done!")

        print("Begin writing AllInOne file...")
        fileIndex = 1
        writeBook = openpyxl.Workbook(write_only=True)
        for fileName in allData:
            if fileName == None:
                continue
            sys.stdout.write("processing file %d/%d \r" %(fileIndex, maxFile))
            sys.stdout.flush()
            fileIndex += 1

            writeSheet = writeBook.create_sheet(title=fileName)
            writeSheet.append(columsName)
            for row in allData[fileName]:
                writeSheet.append(row)
        writeBook.save(outputFileName)
        print("\nWrite AllInOne file Done!")

if __name__ == "__main__":
    print("Begin at %s " % datetime.datetime.now())

    #输入文件
    inputFileName = sys.argv[1]
    
    #原文件存有很多空行，手动定位一下最后一行数据，以便运行时间不会过长
    maxRow = 0 
    
    #输出文件夹
    outputDir = "TextMapFiles_" + inputFileName
    
    #输出的合一文件
    outputFileName = inputFileName + ".AllInOne.xlsx"

    s = Spliter()
    s.Run(inputFileName, maxRow, outputDir, outputFileName)

    print("All Done! at %s " % datetime.datetime.now())